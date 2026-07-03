"""Write the invoice Excel in the required TMC format:
columns = tmc_code | stock_group_code | qty | price
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from .models import PODocument

COLUMNS = ["tmc_code", "stock_group_code", "qty", "price"]


def default_filename(po_no: str) -> str:
    """ชื่อไฟล์ผลลัพธ์: PO Number_วันที่สร้าง.xlsx (วันที่ = วันที่บันทึกไฟล์)."""
    stamp = datetime.now().strftime("%Y%m%d")
    safe_po = "".join(ch for ch in (po_no or "PO") if ch not in '\\/:*?"<>|').strip()
    return f"{safe_po}_{stamp}.xlsx"


def default_output_path(root: str, customer: str, invoice_subfolder: str,
                        po_no: str) -> str:
    """ตำแหน่งเริ่มต้น (โฟลเดอร์ INVOICE FILE ของลูกค้า) — ผู้ใช้เปลี่ยนได้ตอนบันทึก."""
    d = Path(root) / customer / invoice_subfolder
    d.mkdir(parents=True, exist_ok=True)
    return str(d / default_filename(po_no))


def default_combined_filename(customer: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d")
    return f"{customer}_รวม_{stamp}.xlsx"


def default_combined_path(root: str, customer: str, invoice_subfolder: str) -> str:
    d = Path(root) / customer / invoice_subfolder
    d.mkdir(parents=True, exist_ok=True)
    return str(d / default_combined_filename(customer))


def export_combined(docs: list[PODocument], out_path: str) -> str:
    """รวมทุก PO ในโฟลเดอร์เป็นไฟล์เดียว: รายการสินค้าทุกใบต่อกันในชีต 'Invoice' เดียว."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.append(COLUMNS)
    for doc in docs:
        for l in doc.lines:
            ws.append([l.tmc_code, l.stock_group_code, l.qty, l.price])
    # ชีตสรุปแยกรายใบ (ลูกค้า/เลข PO/จำนวน/ยอด) เพื่อการตรวจสอบย้อนหลัง
    s = wb.create_sheet("Summary")
    s.append(["customer", "po_no", "po_date", "item_count", "total", "vat", "grand_total"])
    for doc in docs:
        s.append([doc.customer, doc.po_no, doc.po_date, doc.item_count,
                  doc.total, doc.vat, doc.grand_total])
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def export_invoice(doc: PODocument, out_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.append(COLUMNS)
    for l in doc.lines:
        ws.append([l.tmc_code, l.stock_group_code, l.qty, l.price])
    # a small summary sheet mirrors what the program stores per month
    s = wb.create_sheet("Summary")
    s.append(["customer", "po_no", "po_date", "item_count", "total", "vat", "grand_total"])
    s.append([doc.customer, doc.po_no, doc.po_date, doc.item_count,
              doc.total, doc.vat, doc.grand_total])
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
